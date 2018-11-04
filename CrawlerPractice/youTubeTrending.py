import requests as rq
from bs4 import BeautifulSoup
import pandas as pd
import datetime
from openpyxl import load_workbook


# parse DOM to 2 by 2 array
def _build_data_table(dom_list):
    data_list = []
    for dom in dom_list:
        a_tag_list = dom.find_all('a')
        li_tag_list = dom.find_all('li')

        title = a_tag_list[0]['title']
        link = a_tag_list[0]['href']
        channel = a_tag_list[1].text
        hits = int(li_tag_list[1].text.split('：')[1].split(' ')[0].replace(',', ''))
        start_date = li_tag_list[0].text

        data_list.append([title, 'www.youtube.com' + link, channel, hits, start_date])
    return data_list


# save to Excel with sheet name
def _save_to_excel(data_frame, file_name, sheet_name):
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    try:
        excel_book = load_workbook(file_name)
        writer.book = excel_book
        writer.sheets = dict((ws.title, ws) for ws in excel_book.worksheets)
        data_frame.to_excel(writer, sheet_name)
    except:
        # if file not existed
        data_frame.to_excel(writer, sheet_name)
    finally:
        writer.save()
    _adjust_columns_width(file_name)
    print('save to ' + file_name + ' success.')


# adjust Columns Width
def _adjust_columns_width(file_name):
    mywb = load_workbook(file_name)
    for sheet_name in mywb.sheetnames:
        target_sheet = mywb[sheet_name]
        target_sheet.column_dimensions['A'].width = 5
        target_sheet.column_dimensions['B'].width = 100
        target_sheet.column_dimensions['C'].width = 40
        target_sheet.column_dimensions['D'].width = 15
        target_sheet.column_dimensions['E'].width = 15
    mywb.save(file_name)


def main():
    # crawler
    res = rq.get('https://www.youtube.com/feed/trending')
    soup = BeautifulSoup(res.text, 'lxml')
    dom_list = soup.select('.yt-lockup-content')  # find all link of viedo tag
    data_list = _build_data_table(dom_list)

    # io, save to file
    df = pd.DataFrame(data_list, columns=['Title', 'Link', 'Channel', 'Hits', 'StartDate'])
    file_name = 'YouTubeTrending.xlsx'
    today_str = datetime.date.today().isoformat()
    _save_to_excel(df, file_name, today_str)


if __name__ == '__main__':
    main()